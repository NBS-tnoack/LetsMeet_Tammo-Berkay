from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from pymongo import MongoClient

from db import connect

ROOT = Path(__file__).resolve().parents[1]
EXCEL_SOURCE = ROOT / "Lets Meet DB Dump.xlsx"
SCHEMA = ROOT / "sql" / "03_schema_v2.sql"
HOBBY_PATTERN = re.compile(r"^(.*?) %(-?\d+)%$")


def parse_name(value: object) -> tuple[str, str]:
    parts = str(value).split(", ", 1)
    if len(parts) != 2:
        raise ValueError(f"Ungültiges Namensformat: {value!r}")
    return parts[1], parts[0]


def parse_address(value: object) -> tuple[str, str, str]:
    parts = str(value).split(", ", 2)
    if len(parts) != 3:
        raise ValueError(f"Ungültiges Adressformat: {value!r}")
    return parts


def parse_date(value: object):
    return datetime.strptime(str(value), "%d.%m.%Y").date()


def parse_hobbies(value: object) -> list[tuple[str, int]]:
    # Format in der Excel-Zelle: "Lesen %78%; Kochen %-20%" usw., getrennt durch "; "
    if value is None or str(value) == "":
        return []
    hobbies = []
    for item in str(value).split("; "):
        if not item:
            continue
        match = HOBBY_PATTERN.fullmatch(item)
        if match is None:
            raise ValueError(f"Ungültiges Hobbyformat: {item!r}")
        hobbies.append((match.group(1), int(match.group(2))))
    return hobbies


def read_excel() -> tuple[list[dict[str, object]], dict[str, str]]:
    sheet = load_workbook(EXCEL_SOURCE, read_only=True, data_only=True).active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    people = []
    email_spelling = {}
    for row_number, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        first_name, last_name = parse_name(row["Nachname, Vorname"])
        street, postal_code, city = parse_address(row["Straße Nr, PLZ Ort"])
        email = str(row["E-Mail"])
        record = {
            "email": email,
            "first_name": first_name,
            "last_name": last_name,
            "birth_date": parse_date(row["Geburtsdatum"]),
            "postal_code": postal_code,
            "city": city,
            "phone": None,  # kommt erst aus MongoDB, Excel hat keine Telefonnummer
            "gender": str(row["Geschlecht (m/w/nonbinary)"]),
            "interest_codes": list(str(row["Interessiert an"])),
            "hobbies": parse_hobbies(row["Hobby1 %Prio1%; Hobby2 %Prio2%; Hobby3 %Prio3%; Hobby4 %Prio4%; Hobby5 %Prio5%;"]),
        }
        people.append(record)
        email_spelling[email.casefold()] = email
    if len(email_spelling) != len(people):
        raise ValueError("Excel-E-Mail-Adressen sind nicht eindeutig")
    return people, email_spelling


def parse_timestamp(value: object) -> datetime:
    text = str(value)
    # Mongo liefert die Zeitstempel mal so, mal so - beide Formate kommen tatsaechlich vor
    for pattern in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S"):
        try:
            return datetime.strptime(text, pattern)
        except ValueError:
            continue
    raise ValueError(f"Unbekanntes Zeitformat: {text!r}")


def main() -> None:
    people, email_spelling = read_excel()
    mongo = MongoClient(os.getenv("MONGO_URI", "mongodb://127.0.0.1:27017/"))["LetsMeet"]
    mongo_users = list(mongo.users.find())
    if len(mongo_users) != len(people):
        raise ValueError(f"Erwartet {len(people)} Mongo-Nutzer, gefunden {len(mongo_users)}")
    by_email = {user["_id"].casefold(): user for user in mongo_users}
    if set(by_email) != set(email_spelling):
        raise ValueError("MongoDB- und Excel-E-Mails stimmen nicht überein")
    for person in people:
        person["mongo"] = by_email[person["email"].casefold()]
        person["phone"] = str(person["mongo"].get("phone", ""))
        # MongoDB gilt als der neuere Stand, deshalb ueberschreiben wir den Excel-Namen
        person["first_name"], person["last_name"] = parse_name(person["mongo"].get("name", ""))

    with connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(SCHEMA.read_text(encoding="utf-8"))
            person_ids = {}
            for person in people:
                cursor.execute(
                    """
                    INSERT INTO migration_persons
                        (email, first_name, last_name, birth_date, postal_code, city, phone, gender)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING person_id
                    """,
                    tuple(person[key] for key in ("email", "first_name", "last_name", "birth_date", "postal_code", "city", "phone", "gender")),
                )
                person_ids[person["email"]] = cursor.fetchone()[0]

            interests = []
            hobbies = []
            for person in people:
                person_id = person_ids[person["email"]]
                for code in person["interest_codes"]:
                    interests.append((person_id, code))
                for name, priority in person["hobbies"]:
                    hobbies.append((person_id, name, priority, "excel"))
            likes = []
            messages = []
            for person in people:
                sender_id = person_ids[person["email"]]
                for like in person["mongo"].get("likes", []):
                    target = email_spelling.get(str(like["liked_email"]).casefold())
                    if target is None:
                        raise ValueError(f"Unbekannte Like-Adresse: {like['liked_email']}")
                    likes.append((sender_id, person_ids[target], str(like["status"]), parse_timestamp(like["timestamp"])))
                for message in person["mongo"].get("messages", []):
                    target = email_spelling.get(str(message["receiver_email"]).casefold())
                    if target is None:
                        raise ValueError(f"Unbekannte Nachrichten-Adresse: {message['receiver_email']}")
                    messages.append((sender_id, person_ids[target], str(message["message"]), parse_timestamp(message["timestamp"]), int(message["conversation_id"])))
            for interest in interests:
                cursor.execute(
                    "INSERT INTO migration_interests_data (person_id, interest_code) VALUES (%s, %s)",
                    interest,
                )
            for hobby in hobbies:
                cursor.execute(
                    "INSERT INTO migration_hobbies_data (person_id, hobby_name, priority, source) VALUES (%s, %s, %s, %s)",
                    hobby,
                )
            for like in likes:
                cursor.execute(
                    "INSERT INTO migration_likes_data (liker_id, liked_id, status, liked_at) VALUES (%s, %s, %s, %s)",
                    like,
                )
            for message in messages:
                cursor.execute(
                    "INSERT INTO migration_messages_data (sender_id, receiver_id, body, sent_at, conversation_id) VALUES (%s, %s, %s, %s, %s)",
                    message,
                )
    print(f"V2-Import abgeschlossen: {len(people)} Personen, {len(interests)} Interessen, {len(hobbies)} Hobbys, {len(likes)} Likes, {len(messages)} Nachrichten")


if __name__ == "__main__":
    main()
