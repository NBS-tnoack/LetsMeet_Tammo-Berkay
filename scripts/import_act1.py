from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import execute_values

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Lets Meet DB Dump.xlsx"
SCHEMA = ROOT / "sql" / "01_schema.sql"

HEADERS = {
    "name": "Nachname, Vorname",
    "address": "Straße Nr, PLZ Ort",
    "email": "E-Mail",
    "birth_date": "Geburtsdatum",
}


def parse_row(row: dict[str, object], row_number: int) -> tuple[str, str, str, str, str, str]:
    raw_name = str(row[HEADERS["name"]])
    name_parts = raw_name.split(", ", 1)
    if len(name_parts) != 2:
        raise ValueError(f"Zeile {row_number}: Name ist nicht 'Nachname, Vorname'")
    last_name, first_name = name_parts

    raw_address = str(row[HEADERS["address"]])
    address_parts = raw_address.split(", ", 2)
    if len(address_parts) != 3:
        raise ValueError(f"Zeile {row_number}: Adresse hat nicht drei Teile")
    street, postal_code, city = address_parts

    email = str(row[HEADERS["email"]]) if row[HEADERS["email"]] is not None else ""
    if email == "":
        raise ValueError(f"Zeile {row_number}: E-Mail ist leer")

    raw_birth_date = str(row[HEADERS["birth_date"]])
    try:
        birth_date = datetime.strptime(raw_birth_date, "%d.%m.%Y").date()
    except ValueError as error:
        raise ValueError(f"Zeile {row_number}: ungültiges Geburtsdatum {raw_birth_date!r}") from error

    return email, first_name, last_name, birth_date.isoformat(), postal_code, city


def read_source() -> list[tuple[str, str, str, str, str, str]]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = list(next(rows))
    required = set(HEADERS.values())
    if not required.issubset(headers):
        missing = required.difference(headers)
        raise ValueError(f"Fehlende Spalten: {sorted(missing)}")

    records = []
    for row_number, values in enumerate(rows, start=2):
        source_row = dict(zip(headers, values))
        records.append(parse_row(source_row, row_number))

    emails = [record[0].casefold() for record in records]
    if len(emails) != len(set(emails)):
        raise ValueError("E-Mail-Adressen sind nicht eindeutig (Groß-/Kleinschreibung ignoriert)")
    return records


def connect():
    return psycopg2.connect(
        host=os.getenv("PGHOST", "127.0.0.1"),
        port=os.getenv("PGPORT", "5432"),
        dbname=os.getenv("PGDATABASE", "lf8_lets_meet_db"),
        user=os.getenv("PGUSER", "user"),
        password=os.getenv("PGPASSWORD", "secret"),
    )


def main() -> None:
    records = read_source()
    with connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(SCHEMA.read_text(encoding="utf-8"))
            execute_values(
                cursor,
                """
                INSERT INTO migration_persons
                    (email, first_name, last_name, birth_date, postal_code, city)
                VALUES %s
                """,
                records,
            )
    print(f"Import abgeschlossen: {len(records)} Personen")


if __name__ == "__main__":
    main()
